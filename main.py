#!/usr/bin/env python3

import argparse
import json
import sys
import time
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable, Optional

from openpyxl import load_workbook
from playwright.sync_api import Error as PlaywrightError
from playwright.sync_api import TimeoutError as PlaywrightTimeoutError
from playwright.sync_api import sync_playwright


@dataclass
class ExcelConfig:
    path: Path
    sheet: Optional[str]
    header_row: int
    prompt_column: str
    result_column: str
    start_row: Optional[int]
    skip_completed: bool


@dataclass
class BrowserConfig:
    start_url: str
    user_data_dir: Path
    channel: Optional[str]
    headless: bool
    startup_wait_ms: int
    action_timeout_ms: int


@dataclass
class ChatConfig:
    platform_name: str
    input_selectors: list[str]
    send_button_selectors: list[str]
    response_selectors: list[str]
    new_chat_selectors: list[str]
    loading_selectors: list[str]
    response_timeout_seconds: int
    stability_checks: int
    poll_interval_seconds: float
    send_hotkey: str
    clear_input_hotkey: str
    new_chat_each_prompt: bool
    manual_login: bool


@dataclass
class AppConfig:
    excel: ExcelConfig
    browser: BrowserConfig
    chat: ChatConfig


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Read prompts from Excel, send them to a web chat model, and write results back."
    )
    parser.add_argument(
        "--config",
        default="config.json",
        help="Path to config JSON file. Defaults to ./config.json",
    )
    parser.add_argument(
        "--overwrite",
        action="store_true",
        help="Overwrite existing result cells instead of skipping them.",
    )
    parser.add_argument(
        "--limit",
        type=int,
        default=None,
        help="Only process the first N eligible rows.",
    )
    return parser.parse_args()


def load_config(config_path: Path) -> AppConfig:
    with config_path.open("r", encoding="utf-8") as fh:
        raw = json.load(fh)

    excel_raw = raw["excel"]
    browser_raw = raw["browser"]
    chat_raw = raw.get("chat") or raw.get("doubao")
    if not chat_raw:
        raise ValueError("Config must contain a 'chat' section. Legacy 'doubao' is also supported.")

    excel = ExcelConfig(
        path=(config_path.parent / excel_raw["path"]).expanduser().resolve(),
        sheet=excel_raw.get("sheet"),
        header_row=int(excel_raw.get("header_row", 1)),
        prompt_column=excel_raw["prompt_column"],
        result_column=excel_raw["result_column"],
        start_row=excel_raw.get("start_row"),
        skip_completed=bool(excel_raw.get("skip_completed", True)),
    )
    browser = BrowserConfig(
        start_url=browser_raw.get("start_url") or browser_raw["doubao_url"],
        user_data_dir=(config_path.parent / browser_raw["user_data_dir"]).expanduser().resolve(),
        channel=browser_raw.get("channel"),
        headless=bool(browser_raw.get("headless", False)),
        startup_wait_ms=int(browser_raw.get("startup_wait_ms", 2000)),
        action_timeout_ms=int(browser_raw.get("action_timeout_ms", 15000)),
    )
    chat = ChatConfig(
        platform_name=chat_raw.get("platform_name", "web-chat"),
        input_selectors=list(chat_raw["input_selectors"]),
        send_button_selectors=list(chat_raw.get("send_button_selectors", [])),
        response_selectors=list(chat_raw["response_selectors"]),
        new_chat_selectors=list(chat_raw.get("new_chat_selectors", [])),
        loading_selectors=list(chat_raw.get("loading_selectors", [])),
        response_timeout_seconds=int(chat_raw.get("response_timeout_seconds", 120)),
        stability_checks=int(chat_raw.get("stability_checks", 3)),
        poll_interval_seconds=float(chat_raw.get("poll_interval_seconds", 1.0)),
        send_hotkey=chat_raw.get("send_hotkey", "Enter"),
        clear_input_hotkey=chat_raw.get("clear_input_hotkey", "Meta+A"),
        new_chat_each_prompt=bool(chat_raw.get("new_chat_each_prompt", False)),
        manual_login=bool(chat_raw.get("manual_login", True)),
    )
    return AppConfig(excel=excel, browser=browser, chat=chat)


def column_index_by_header(sheet, header_row: int, header_name: str) -> int:
    normalized_target = str(header_name).strip().lower()
    for cell in sheet[header_row]:
        if str(cell.value or "").strip().lower() == normalized_target:
            return cell.column
    raise ValueError(f"Header '{header_name}' not found in row {header_row}")


def iter_candidate_rows(sheet, excel_config: ExcelConfig, overwrite: bool) -> Iterable[int]:
    prompt_col = column_index_by_header(sheet, excel_config.header_row, excel_config.prompt_column)
    result_col = column_index_by_header(sheet, excel_config.header_row, excel_config.result_column)
    start_row = excel_config.start_row or (excel_config.header_row + 1)

    for row_idx in range(start_row, sheet.max_row + 1):
        prompt_value = sheet.cell(row=row_idx, column=prompt_col).value
        result_value = sheet.cell(row=row_idx, column=result_col).value
        prompt_text = str(prompt_value).strip() if prompt_value is not None else ""
        result_text = str(result_value).strip() if result_value is not None else ""
        if not prompt_text:
            continue
        if excel_config.skip_completed and result_text and not overwrite:
            continue
        yield row_idx


def resolve_first_locator(page, selectors: list[str], require_visible: bool) -> Optional[tuple[str, object]]:
    for selector in selectors:
        locator = page.locator(selector)
        if locator.count() == 0:
            continue
        if not require_visible:
            return selector, locator.first
        try:
            if locator.first.is_visible():
                return selector, locator.first
        except PlaywrightError:
            continue
    return None


def ensure_chat_ready(page, config: AppConfig) -> tuple[str, object]:
    deadline = time.time() + max(30, config.chat.response_timeout_seconds)
    while time.time() < deadline:
        resolved = resolve_first_locator(page, config.chat.input_selectors, require_visible=True)
        if resolved:
            return resolved
        time.sleep(config.chat.poll_interval_seconds)
    raise RuntimeError(
        f"Could not find a visible input box for {config.chat.platform_name}. "
        "Update chat.input_selectors in config.json."
    )


def click_if_present(page, selectors: list[str], action_timeout_ms: int) -> bool:
    resolved = resolve_first_locator(page, selectors, require_visible=True)
    if not resolved:
        return False
    _, locator = resolved
    try:
        locator.click(timeout=action_timeout_ms)
        return True
    except PlaywrightError:
        return False


def get_last_response_text(page, selectors: list[str]) -> str:
    for selector in selectors:
        locator = page.locator(selector)
        count = locator.count()
        if count == 0:
            continue
        text = locator.nth(count - 1).inner_text().strip()
        if text:
            return text
    return ""


def get_response_count(page, selectors: list[str]) -> int:
    for selector in selectors:
        count = page.locator(selector).count()
        if count > 0:
            return count
    return 0


def is_loading(page, selectors: list[str]) -> bool:
    for selector in selectors:
        locator = page.locator(selector)
        if locator.count() == 0:
            continue
        try:
            if locator.first.is_visible():
                return True
        except PlaywrightError:
            continue
    return False


def prepare_input(locator, prompt: str, clear_input_hotkey: str) -> None:
    locator.click()
    try:
        locator.fill("")
        locator.fill(prompt)
        return
    except PlaywrightError:
        pass

    locator.press(clear_input_hotkey)
    locator.press("Backspace")
    locator.type(prompt, delay=20)


def wait_for_response(page, config: AppConfig, previous_count: int, previous_text: str) -> str:
    deadline = time.time() + config.chat.response_timeout_seconds
    latest_text = previous_text
    stable_rounds = 0

    while time.time() < deadline:
        current_count = get_response_count(page, config.chat.response_selectors)
        current_text = get_last_response_text(page, config.chat.response_selectors)
        response_arrived = current_count > previous_count or (
            current_text and current_text != previous_text
        )

        if response_arrived and current_text:
            if current_text == latest_text:
                stable_rounds += 1
            else:
                latest_text = current_text
                stable_rounds = 0

            if stable_rounds >= config.chat.stability_checks and not is_loading(
                page, config.chat.loading_selectors
            ):
                return current_text

        time.sleep(config.chat.poll_interval_seconds)

    raise TimeoutError(f"Timed out waiting for {config.chat.platform_name} to finish responding.")


def send_prompt(page, config: AppConfig, prompt: str) -> str:
    if config.chat.new_chat_each_prompt:
        click_if_present(page, config.chat.new_chat_selectors, config.browser.action_timeout_ms)
        time.sleep(1)

    _, input_locator = ensure_chat_ready(page, config)
    previous_count = get_response_count(page, config.chat.response_selectors)
    previous_text = get_last_response_text(page, config.chat.response_selectors)

    prepare_input(input_locator, prompt, config.chat.clear_input_hotkey)
    if not click_if_present(page, config.chat.send_button_selectors, config.browser.action_timeout_ms):
        input_locator.press(config.chat.send_hotkey)

    return wait_for_response(page, config, previous_count, previous_text)


def main() -> int:
    args = parse_args()
    config_path = Path(args.config).expanduser().resolve()
    config = load_config(config_path)

    workbook = load_workbook(config.excel.path)
    sheet = workbook[config.excel.sheet] if config.excel.sheet else workbook.active
    prompt_col = column_index_by_header(sheet, config.excel.header_row, config.excel.prompt_column)
    result_col = column_index_by_header(sheet, config.excel.header_row, config.excel.result_column)

    rows = list(iter_candidate_rows(sheet, config.excel, overwrite=args.overwrite))
    if args.limit is not None:
        rows = rows[: args.limit]

    if not rows:
        print("No eligible rows found. Nothing to do.")
        return 0

    config.browser.user_data_dir.mkdir(parents=True, exist_ok=True)

    with sync_playwright() as playwright:
        launch_kwargs = {
            "user_data_dir": str(config.browser.user_data_dir),
            "headless": config.browser.headless,
        }
        if config.browser.channel:
            launch_kwargs["channel"] = config.browser.channel

        context = playwright.chromium.launch_persistent_context(**launch_kwargs)
        context.set_default_timeout(config.browser.action_timeout_ms)
        page = context.pages[0] if context.pages else context.new_page()
        page.goto(config.browser.start_url, wait_until="domcontentloaded")
        page.wait_for_timeout(config.browser.startup_wait_ms)

        if config.chat.manual_login:
            try:
                ensure_chat_ready(page, config)
            except RuntimeError:
                print(
                    f"Please log in to {config.chat.platform_name} in the opened browser window, "
                    "then press Enter here."
                )
                input()
                ensure_chat_ready(page, config)

        for index, row_idx in enumerate(rows, start=1):
            prompt = str(sheet.cell(row=row_idx, column=prompt_col).value).strip()
            print(f"[{index}/{len(rows)}] Processing row {row_idx}")
            try:
                result = send_prompt(page, config, prompt)
            except (PlaywrightError, PlaywrightTimeoutError, TimeoutError) as exc:
                result = f"ERROR: {exc}"
            sheet.cell(row=row_idx, column=result_col).value = result
            workbook.save(config.excel.path)

        context.close()

    print(f"Finished. Results written to {config.excel.path}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
