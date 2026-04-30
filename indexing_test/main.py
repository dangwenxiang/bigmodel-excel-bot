#!/usr/bin/env python3

import argparse
import json
import re
import sys
import time
from contextlib import contextmanager
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Iterable, Optional

from openpyxl import Workbook, load_workbook
from playwright.sync_api import Error as PlaywrightError
from playwright.sync_api import TimeoutError as PlaywrightTimeoutError
from playwright.sync_api import sync_playwright


@dataclass
class ExcelConfig:
    path: Path
    sheet: Optional[str]
    header_row: int
    prompt_column: str
    result_column: str
    source_column: Optional[str]
    start_row: Optional[int]
    skip_completed: bool


@dataclass
class BrowserConfig:
    start_url: str
    user_data_dir: Path
    channel: Optional[str]
    headless: bool
    startup_wait_ms: int
    action_timeout_ms: int


@dataclass
class ChatConfig:
    platform_name: str
    input_selectors: list[str]
    send_button_selectors: list[str]
    response_selectors: list[str]
    new_chat_selectors: list[str]
    loading_selectors: list[str]
    popup_selectors: list[str]
    popup_confirm_selectors: list[str]
    popup_artifact_dir: Path
    response_timeout_seconds: int
    stability_checks: int
    poll_interval_seconds: float
    send_hotkey: str
    clear_input_hotkey: str
    new_chat_each_prompt: bool
    manual_login: bool
    manual_popup_confirmation: bool


@dataclass
class AppConfig:
    excel: ExcelConfig
    browser: BrowserConfig
    chat: ChatConfig


@dataclass
class ResponseData:
    text: str
    sources: str
    source_urls: str = ""
    source_titles: str = ""


@dataclass
class SourceRecord:
    url: str
    title: str
    app_name: str


@dataclass
class PromptRunRecord:
    query: str
    result: str
    sources: str
    source_urls: str
    source_titles: str


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Read prompts from Excel, send them to a web chat model, and write results back."
    )
    parser.add_argument(
        "--config",
        default="config.json",
        help="Path to config JSON file. Defaults to ./config.json",
    )
    parser.add_argument(
        "--overwrite",
        action="store_true",
        help="Overwrite existing result cells instead of skipping them.",
    )
    parser.add_argument(
        "--limit",
        type=int,
        default=None,
        help="Only process the first N eligible rows.",
    )
    return parser.parse_args()


def load_config(config_path: Path) -> AppConfig:
    # Accept UTF-8 BOM because many Windows editors emit it by default.
    with config_path.open("r", encoding="utf-8-sig") as fh:
        raw = json.load(fh)

    excel_raw = raw["excel"]
    browser_raw = raw["browser"]
    chat_raw = raw.get("chat") or raw.get("doubao")
    if not chat_raw:
        raise ValueError("Config must contain a 'chat' section. Legacy 'doubao' is also supported.")

    start_url = browser_raw.get("start_url") or browser_raw["doubao_url"]
    start_url_path = (config_path.parent / start_url).expanduser()
    if "://" not in start_url and start_url_path.exists():
        start_url = start_url_path.resolve().as_uri()

    excel = ExcelConfig(
        path=(config_path.parent / excel_raw["path"]).expanduser().resolve(),
        sheet=excel_raw.get("sheet"),
        header_row=int(excel_raw.get("header_row", 1)),
        prompt_column=excel_raw["prompt_column"],
        result_column=excel_raw["result_column"],
        source_column=excel_raw.get("source_column"),
        start_row=excel_raw.get("start_row"),
        skip_completed=bool(excel_raw.get("skip_completed", True)),
    )
    browser = BrowserConfig(
        start_url=start_url,
        user_data_dir=(config_path.parent / browser_raw["user_data_dir"]).expanduser().resolve(),
        channel=browser_raw.get("channel"),
        headless=bool(browser_raw.get("headless", False)),
        startup_wait_ms=int(browser_raw.get("startup_wait_ms", 2000)),
        action_timeout_ms=int(browser_raw.get("action_timeout_ms", 15000)),
    )
    chat = ChatConfig(
        platform_name=chat_raw.get("platform_name", "web-chat"),
        input_selectors=list(chat_raw["input_selectors"]),
        send_button_selectors=list(chat_raw.get("send_button_selectors", [])),
        response_selectors=list(chat_raw["response_selectors"]),
        new_chat_selectors=list(chat_raw.get("new_chat_selectors", [])),
        loading_selectors=list(chat_raw.get("loading_selectors", [])),
        popup_selectors=list(chat_raw.get("popup_selectors", ["[role='dialog']", "[aria-modal='true']"])),
        popup_confirm_selectors=list(chat_raw.get("popup_confirm_selectors", [])),
        popup_artifact_dir=(config_path.parent / chat_raw.get("popup_artifact_dir", "./artifacts/popups"))
        .expanduser()
        .resolve(),
        response_timeout_seconds=int(chat_raw.get("response_timeout_seconds", 120)),
        stability_checks=int(chat_raw.get("stability_checks", 3)),
        poll_interval_seconds=float(chat_raw.get("poll_interval_seconds", 1.0)),
        send_hotkey=chat_raw.get("send_hotkey", "Enter"),
        clear_input_hotkey=chat_raw.get("clear_input_hotkey", "Control+A"),
        new_chat_each_prompt=bool(chat_raw.get("new_chat_each_prompt", False)),
        manual_login=bool(chat_raw.get("manual_login", True)),
        manual_popup_confirmation=bool(chat_raw.get("manual_popup_confirmation", True)),
    )
    return AppConfig(excel=excel, browser=browser, chat=chat)


_POPUP_CAPTURED_STATE: dict[int, bool] = {}
_POPUP_COUNTER_STATE: dict[int, int] = {}


def column_index_by_header(sheet, header_row: int, header_name: str) -> int:
    normalized_target = str(header_name).strip().lower()
    for cell in sheet[header_row]:
        if str(cell.value or "").strip().lower() == normalized_target:
            return cell.column
    raise ValueError(f"Header '{header_name}' not found in row {header_row}")


def ensure_column_index_by_header(sheet, header_row: int, header_name: Optional[str]) -> Optional[int]:
    if not header_name:
        return None
    try:
        return column_index_by_header(sheet, header_row, header_name)
    except ValueError:
        new_column = sheet.max_column + 1
        sheet.cell(row=header_row, column=new_column).value = header_name
        return new_column


def iter_candidate_rows(sheet, excel_config: ExcelConfig, overwrite: bool) -> Iterable[int]:
    prompt_col = column_index_by_header(sheet, excel_config.header_row, excel_config.prompt_column)
    result_col = column_index_by_header(sheet, excel_config.header_row, excel_config.result_column)
    start_row = excel_config.start_row or (excel_config.header_row + 1)

    for row_idx in range(start_row, sheet.max_row + 1):
        prompt_value = sheet.cell(row=row_idx, column=prompt_col).value
        result_value = sheet.cell(row=row_idx, column=result_col).value
        prompt_text = str(prompt_value).strip() if prompt_value is not None else ""
        result_text = str(result_value).strip() if result_value is not None else ""
        if not prompt_text:
            continue
        if excel_config.skip_completed and result_text and not overwrite:
            continue
        yield row_idx


def resolve_first_locator(page, selectors: list[str], require_visible: bool) -> Optional[tuple[str, object]]:
    for selector in selectors:
        locator = page.locator(selector)
        if locator.count() == 0:
            continue
        if not require_visible:
            return selector, locator.first
        try:
            if locator.first.is_visible():
                return selector, locator.first
        except PlaywrightError:
            continue
    return None


def ensure_chat_ready(page, config: AppConfig) -> tuple[str, object]:
    deadline = time.time() + max(30, config.chat.response_timeout_seconds)
    while time.time() < deadline:
        resolved = resolve_first_locator(page, config.chat.input_selectors, require_visible=True)
        if resolved:
            return resolved
        time.sleep(config.chat.poll_interval_seconds)
    raise RuntimeError(
        f"Could not find a visible input box for {config.chat.platform_name}. "
        "Update chat.input_selectors in config.json."
    )


def click_if_present(page, selectors: list[str], action_timeout_ms: int) -> bool:
    resolved = resolve_first_locator(page, selectors, require_visible=True)
    if not resolved:
        return False
    _, locator = resolved
    try:
        locator.click(timeout=action_timeout_ms)
        return True
    except PlaywrightError:
        return False


def save_popup_artifacts(page, config: AppConfig) -> Path:
    page_key = id(page)
    popup_index = _POPUP_COUNTER_STATE.get(page_key, 0) + 1
    _POPUP_COUNTER_STATE[page_key] = popup_index

    artifact_dir = (
        config.chat.popup_artifact_dir
        / f"{config.chat.platform_name}-{datetime.now().strftime('%Y%m%d-%H%M%S')}-{popup_index:02d}"
    )
    artifact_dir.mkdir(parents=True, exist_ok=True)

    popup_entries = []
    for selector in config.chat.popup_selectors:
        locator = page.locator(selector)
        count = locator.count()
        for index in range(count):
            popup = locator.nth(index)
            try:
                if not popup.is_visible():
                    continue
                popup_entries.append(
                    {
                        "selector": selector,
                        "index": index,
                        "text": popup.inner_text(),
                        "html": popup.evaluate("(el) => el.outerHTML"),
                    }
                )
            except PlaywrightError:
                continue

    button_entries = []
    buttons = page.locator("button")
    for index in range(buttons.count()):
        button = buttons.nth(index)
        try:
            if not button.is_visible():
                continue
            button_entries.append(
                {
                    "index": index,
                    "text": button.inner_text().strip(),
                    "aria_label": button.get_attribute("aria-label"),
                    "class": button.get_attribute("class"),
                }
            )
        except PlaywrightError:
            continue

    (artifact_dir / "popup-dom.json").write_text(
        json.dumps(
            {
                "captured_at": datetime.now().isoformat(),
                "page_url": page.url,
                "page_title": page.title(),
                "popup_selectors": config.chat.popup_selectors,
                "popup_confirm_selectors": config.chat.popup_confirm_selectors,
                "popups": popup_entries,
                "visible_buttons": button_entries,
            },
            ensure_ascii=False,
            indent=2,
        ),
        encoding="utf-8",
    )
    page.screenshot(path=str(artifact_dir / "page.png"), full_page=True)
    for index, entry in enumerate(popup_entries, start=1):
        (artifact_dir / f"popup-{index:02d}.html").write_text(entry["html"], encoding="utf-8")

    return artifact_dir


def detect_popup(page, selectors: list[str]) -> bool:
    for selector in selectors:
        locator = page.locator(selector)
        if locator.count() == 0:
            continue
        try:
            if locator.first.is_visible():
                return True
        except PlaywrightError:
            continue
    return False


def handle_popup_if_present(page, config: AppConfig) -> bool:
    page_key = id(page)
    if not detect_popup(page, config.chat.popup_selectors):
        _POPUP_CAPTURED_STATE[page_key] = False
        return False

    if not _POPUP_CAPTURED_STATE.get(page_key, False):
        artifact_dir = save_popup_artifacts(page, config)
        _POPUP_CAPTURED_STATE[page_key] = True
        print(f"Popup artifacts saved to {artifact_dir}")

    if click_if_present(page, config.chat.popup_confirm_selectors, config.browser.action_timeout_ms):
        page.wait_for_timeout(1000)
        if not detect_popup(page, config.chat.popup_selectors):
            _POPUP_CAPTURED_STATE[page_key] = False
        return True

    if config.chat.manual_popup_confirmation:
        print(
            f"Detected a popup in {config.chat.platform_name}. "
            "Complete the verification in the browser. The script will continue automatically."
        )
        deadline = time.time() + max(60, config.chat.response_timeout_seconds)
        while time.time() < deadline:
            if not detect_popup(page, config.chat.popup_selectors):
                _POPUP_CAPTURED_STATE[page_key] = False
                page.wait_for_timeout(1000)
                return True
            time.sleep(config.chat.poll_interval_seconds)
        raise TimeoutError(
            f"Timed out waiting for popup verification in {config.chat.platform_name}."
        )
        return True

    return False


def extract_reference_section(text: str) -> str:
    lines = [line.rstrip() for line in text.splitlines()]
    heading_patterns = (
        "参考资料",
        "参考来源",
        "资料来源",
        "参考链接",
        "引用来源",
        "sources",
        "references",
    )
    collecting = False
    collected_lines: list[str] = []

    for line in lines:
        stripped = line.strip()
        normalized = stripped.lower().rstrip(":：")
        if not collecting and normalized in heading_patterns:
            collecting = True
            continue
        if not collecting:
            continue
        if not stripped:
            if collected_lines:
                break
            continue
        collected_lines.append(stripped)
        if len(collected_lines) >= 12:
            break

    return "\n".join(collected_lines).strip()


def format_sources(text: str, links: list[dict[str, str]]) -> str:
    items: list[str] = []
    seen: set[str] = set()
    seen_urls: set[str] = set()
    url_pattern = re.compile(r"https?://[^\s)\]}>]+")

    reference_section = extract_reference_section(text)
    if reference_section:
        for line in reference_section.splitlines():
            normalized = line.strip()
            if normalized and normalized not in seen:
                seen.add(normalized)
                items.append(normalized)
                seen_urls.update(url_pattern.findall(normalized))

    for link in links:
        href = (link.get("href") or "").strip()
        label = " ".join((link.get("text") or "").split())
        candidate = f"{label} {href}".strip() if label else href
        if href and href not in seen_urls and candidate not in seen:
            seen.add(candidate)
            items.append(candidate)
            seen_urls.add(href)

    for url in url_pattern.findall(text):
        if url not in seen and url not in seen_urls:
            seen.add(url)
            items.append(url)
            seen_urls.add(url)

    return "\n".join(items)


def normalize_source_url(url: str) -> str:
    return url.strip().rstrip("/")


def dedupe_source_records(records: list[SourceRecord]) -> list[SourceRecord]:
    deduped: list[SourceRecord] = []
    seen_keys: set[tuple[str, str, str]] = set()

    for record in records:
        normalized = SourceRecord(
            url=(record.url or "").strip(),
            title=(record.title or "").strip(),
            app_name=(record.app_name or "").strip(),
        )
        if not any((normalized.url, normalized.title, normalized.app_name)):
            continue
        key = (
            normalize_source_url(normalized.url),
            normalized.title,
            normalized.app_name,
        )
        if key in seen_keys:
            continue
        seen_keys.add(key)
        deduped.append(normalized)

    return deduped


def build_structured_source_fields(records: list[SourceRecord]) -> tuple[str, str, str]:
    records = dedupe_source_records(records)
    sources = "\n".join(
        dict.fromkeys(record.app_name for record in records if record.app_name).keys()
    )
    source_urls = "\n".join(
        dict.fromkeys(record.url for record in records if record.url).keys()
    )
    source_titles = "\n".join(
        dict.fromkeys(record.title for record in records if record.title).keys()
    )
    return sources, source_urls, source_titles


def get_nearby_dom_source_records(page, selectors: list[str]) -> list[SourceRecord]:
    try:
        payload = page.evaluate(
            r"""(selectors) => {
            const selectorText = selectors.join(',');
            const responseNodes = Array.from(document.querySelectorAll(selectorText));
            const responseNode = responseNodes.length ? responseNodes[responseNodes.length - 1] : null;
            if (!responseNode) {
                return [];
            }

            const normalize = (value) => String(value || '').replace(/\s+/g, ' ').trim();
            const isUsefulHref = (href) => {
                const value = String(href || '').trim();
                if (!value || value.startsWith('javascript:') || value.startsWith('data:')) {
                    return false;
                }
                try {
                    const url = new URL(value, location.href);
                    if (url.hostname.endsWith('doubao.com') && url.pathname.startsWith('/chat')) {
                        return false;
                    }
                    return true;
                } catch {
                    return false;
                }
            };
            const toRecord = (link) => {
                const href = link.href || link.getAttribute('href') || '';
                let hostname = '';
                try {
                    hostname = new URL(href, location.href).hostname.replace(/^www\./, '');
                } catch {}
                return {
                    url: href,
                    title: normalize(link.innerText || link.getAttribute('title') || link.getAttribute('aria-label') || ''),
                    app_name: normalize(link.getAttribute('data-site-name') || link.getAttribute('data-source') || hostname),
                };
            };

            let current = responseNode;
            for (let depth = 0; current && depth < 10; depth += 1, current = current.parentElement) {
                const links = Array.from(current.querySelectorAll('a[href]'))
                    .filter((link) => isUsefulHref(link.href || link.getAttribute('href')))
                    .map(toRecord)
                    .filter((item) => item.url || item.title || item.app_name);
                if (links.length > 0) {
                    return links.slice(0, 50);
                }
            }
            return [];
        }""",
            selectors,
        )
    except PlaywrightError:
        return []

    if not isinstance(payload, list):
        return []
    records: list[SourceRecord] = []
    for item in payload:
        if not isinstance(item, dict):
            continue
        records.append(
            SourceRecord(
                url=str(item.get("url") or "").strip(),
                title=str(item.get("title") or "").strip(),
                app_name=str(item.get("app_name") or "").strip(),
            )
        )
    return dedupe_source_records(records)


def get_reference_panel_source_records(page) -> list[SourceRecord]:
    try:
        page.wait_for_function(
            r"""() => Array.from(document.querySelectorAll('div,span,button,[role="button"]'))
                .some((el) => {
                    const text = String(el.innerText || el.textContent || '').replace(/\s+/g, ' ').trim();
                    const rect = el.getBoundingClientRect();
                    return /^参考\s*\d+\s*篇资料$/.test(text) && rect.width > 0 && rect.height > 0;
                })""",
            timeout=5000,
        )
        clicked = page.evaluate(
            r"""() => {
            const normalize = (value) => String(value || '').replace(/\s+/g, ' ').trim();
            const isVisible = (el) => {
                const rect = el.getBoundingClientRect();
                const style = window.getComputedStyle(el);
                return rect.width > 0 && rect.height > 0 && style.visibility !== 'hidden' && style.display !== 'none';
            };
            const looksClickable = (el) => {
                if (!el || !isVisible(el)) {
                    return false;
                }
                const role = el.getAttribute('role') || '';
                const cls = el.getAttribute('class') || '';
                const style = window.getComputedStyle(el);
                return Boolean(
                    el.tagName === 'BUTTON' ||
                    el.tagName === 'A' ||
                    role === 'button' ||
                    el.hasAttribute('onclick') ||
                    el.hasAttribute('tabindex') ||
                    style.cursor === 'pointer' ||
                    /reference|ref|citation|source|search/i.test(cls)
                );
            };
            const clickableTarget = (el) => {
                let current = el;
                for (let depth = 0; current && depth < 6; depth += 1, current = current.parentElement) {
                    if (looksClickable(current)) {
                        return current;
                    }
                }
                return el;
            };
            const candidates = Array.from(document.querySelectorAll('div,span,button,[role="button"]'))
                .filter((el) => /^参考\s*\d+\s*篇资料$/.test(normalize(el.innerText || el.textContent || '')))
                .filter(isVisible);
            const target = candidates.length ? candidates[candidates.length - 1] : null;
            if (!target) {
                return false;
            }
            const actionTarget = clickableTarget(target);
            actionTarget.scrollIntoView({ block: 'center', inline: 'center' });
            const rect = actionTarget.getBoundingClientRect();
            const eventOptions = {
                bubbles: true,
                cancelable: true,
                view: window,
                clientX: rect.left + rect.width / 2,
                clientY: rect.top + rect.height / 2,
            };
            for (const type of ['pointerdown', 'mousedown', 'mouseup', 'click']) {
                actionTarget.dispatchEvent(new MouseEvent(type, eventOptions));
            }
            if (typeof actionTarget.click === 'function') {
                actionTarget.click();
            }
            return true;
        }"""
        )
        if not clicked:
            return []
        page.wait_for_function(
            r"""() => Array.from(document.querySelectorAll('a[href]'))
                .some((link) => {
                    const rect = link.getBoundingClientRect();
                    const href = String(link.href || link.getAttribute('href') || '').trim();
                    if (rect.width <= 0 || rect.height <= 0 || !href || href.startsWith('javascript:') || href.startsWith('data:')) {
                        return false;
                    }
                    try {
                        const url = new URL(href, location.href);
                        return !(url.hostname.endsWith('doubao.com') && url.pathname.startsWith('/chat'));
                    } catch {
                        return false;
                    }
                })""",
            timeout=5000,
        )
        payload = page.evaluate(
            r"""() => {
            const normalize = (value) => String(value || '').replace(/\s+/g, ' ').trim();
            const isUsefulHref = (href) => {
                const value = String(href || '').trim();
                if (!value || value.startsWith('javascript:') || value.startsWith('data:')) {
                    return false;
                }
                try {
                    const url = new URL(value, location.href);
                    if (url.hostname.endsWith('doubao.com') && url.pathname.startsWith('/chat')) {
                        return false;
                    }
                    return true;
                } catch {
                    return false;
                }
            };
            return Array.from(document.querySelectorAll('a[href]'))
                .filter((link) => {
                    const rect = link.getBoundingClientRect();
                    return rect.width > 0 && rect.height > 0 && isUsefulHref(link.href || link.getAttribute('href'));
                })
                .map((link) => {
                    let hostname = '';
                    try {
                        hostname = new URL(link.href || link.getAttribute('href') || '', location.href).hostname.replace(/^www\./, '');
                    } catch {}
                    const text = normalize(link.innerText || link.textContent || '');
                    const title = normalize(link.getAttribute('title') || link.getAttribute('aria-label') || '');
                    return {
                        url: link.href || link.getAttribute('href') || '',
                        title: text || title,
                        app_name: normalize(link.getAttribute('data-site-name') || link.getAttribute('data-source') || hostname),
                    };
                })
                .filter((item) => item.url || item.title || item.app_name)
                .slice(0, 50);
        }"""
        )
    except PlaywrightError:
        return []

    if not isinstance(payload, list):
        return []
    records: list[SourceRecord] = []
    for item in payload:
        if not isinstance(item, dict):
            continue
        records.append(
            SourceRecord(
                url=str(item.get("url") or "").strip(),
                title=str(item.get("title") or "").strip(),
                app_name=str(item.get("app_name") or "").strip(),
            )
        )
    return dedupe_source_records(records)


def save_source_debug_artifact(
    page,
    config: AppConfig,
    expected_prompt: str,
    expected_response_text: str,
    links: list[dict[str, str]],
) -> Optional[Path]:
    try:
        payload = page.evaluate(
            r"""(args) => {
            const keywords = ['source', 'sources', 'reference', 'references', 'search', 'citation', 'url', 'title'];
            const maxMatches = 300;
            const maxDepth = 10;
            const seen = new WeakSet();
            const matches = [];

            const short = (value) => {
                if (value === null || value === undefined) {
                    return value;
                }
                const text = String(value);
                return text.length > 500 ? `${text.slice(0, 500)}...` : text;
            };

            const isInterestingKey = (key) => {
                const lower = String(key || '').toLowerCase();
                return keywords.some((keyword) => lower.includes(keyword));
            };

            const walk = (value, path, depth) => {
                if (matches.length >= maxMatches || depth > maxDepth || value === null || value === undefined) {
                    return;
                }
                if (typeof value !== 'object') {
                    return;
                }
                if (seen.has(value)) {
                    return;
                }
                seen.add(value);

                if (Array.isArray(value)) {
                    for (let index = 0; index < Math.min(value.length, 50); index += 1) {
                        walk(value[index], `${path}[${index}]`, depth + 1);
                    }
                    return;
                }

                for (const [key, item] of Object.entries(value)) {
                    const itemPath = path ? `${path}.${key}` : key;
                    if (isInterestingKey(key)) {
                        matches.push({
                            path: itemPath,
                            type: Array.isArray(item) ? 'array' : typeof item,
                            value: typeof item === 'object' ? short(JSON.stringify(item)) : short(item),
                        });
                        if (matches.length >= maxMatches) {
                            return;
                        }
                    }
                    walk(item, itemPath, depth + 1);
                }
            };

            const responseLinks = Array.from(document.querySelectorAll('a[href]')).map((link) => ({
                text: (link.innerText || '').trim(),
                href: link.href || '',
                aria_label: link.getAttribute('aria-label') || '',
                title: link.getAttribute('title') || '',
                class: link.getAttribute('class') || '',
            })).slice(-200);

            const responseNodes = Array.from(document.querySelectorAll(args.responseSelectors.join(','))).map((el, index) => ({
                index,
                text_sample: (el.innerText || '').trim().slice(0, 800),
                links: Array.from(el.querySelectorAll('a[href]')).map((link) => ({
                    text: (link.innerText || '').trim(),
                    href: link.href || '',
                    aria_label: link.getAttribute('aria-label') || '',
                    title: link.getAttribute('title') || '',
                })),
            })).slice(-10);
            const lastResponseNode = responseNodes.length ? document.querySelectorAll(args.responseSelectors.join(','))[document.querySelectorAll(args.responseSelectors.join(',')).length - 1] : null;
            const ancestor_nodes = [];
            let current = lastResponseNode;
            for (let depth = 0; current && depth < 10; depth += 1, current = current.parentElement) {
                ancestor_nodes.push({
                    depth,
                    tag: current.tagName,
                    class: current.getAttribute('class') || '',
                    text_sample: (current.innerText || '').trim().slice(0, 1000),
                    links: Array.from(current.querySelectorAll('a[href]')).map((link) => ({
                        text: (link.innerText || '').trim(),
                        href: link.href || '',
                        aria_label: link.getAttribute('aria-label') || '',
                        title: link.getAttribute('title') || '',
                        class: link.getAttribute('class') || '',
                    })).slice(0, 100),
                });
            }

            const routerData = window._ROUTER_DATA || null;
            walk(routerData, 'window._ROUTER_DATA', 0);

            return {
                page_url: location.href,
                page_title: document.title,
                expected_prompt: args.expectedPrompt,
                expected_response_sample: String(args.expectedResponseText || '').slice(0, 1000),
                router_data_exists: Boolean(routerData),
                router_data_top_keys: routerData && typeof routerData === 'object' ? Object.keys(routerData) : [],
                interesting_router_matches: matches,
                all_page_links_tail: responseLinks,
                response_nodes: responseNodes,
                ancestor_nodes,
            };
        }""",
            {
                "expectedPrompt": expected_prompt,
                "expectedResponseText": expected_response_text,
                "responseSelectors": config.chat.response_selectors,
            },
        )
    except PlaywrightError as exc:
        payload = {"error": str(exc)}

    debug_dir = config.chat.popup_artifact_dir.parent / "source-debug"
    debug_dir.mkdir(parents=True, exist_ok=True)
    output_path = debug_dir / f"{config.chat.platform_name}-{datetime.now().strftime('%Y%m%d-%H%M%S-%f')}.json"
    output_path.write_text(
        json.dumps(
            {
                "captured_at": datetime.now().isoformat(),
                "expected_prompt": expected_prompt,
                "expected_response_sample": expected_response_text[:1000],
                "dom_links_from_response_node": links,
                "page_probe": payload,
            },
            ensure_ascii=False,
            indent=2,
        ),
        encoding="utf-8",
    )
    return output_path


def _extract_structured_source_records_from_loaded_page(
    page,
    expected_prompt: str = "",
    expected_response_text: str = "",
) -> list[SourceRecord]:
    try:
        payload = page.evaluate(
            r"""(args) => {
            const expectedPrompt = args?.expectedPrompt || '';
            const expectedResponseText = args?.expectedResponseText || '';
            const routerData = window._ROUTER_DATA;
            const chatLayout = routerData?.loaderData?.chat_layout;
            const cells = chatLayout?.trimmedChainRecentConvCells ||
                chatLayout?.chat_layout?.trimmedChainRecentConvCells;
            if (!Array.isArray(cells)) {
                return [];
            }

            const activeConversationId = (() => {
                const match = String(globalThis.location?.pathname || '').match(/\/chat\/([^/?#]+)/);
                return match ? match[1] : '';
            })();
            const promptText = String(expectedPrompt || '').trim();
            const responseText = String(expectedResponseText || '').trim();
            const normalizeUrl = (value) => String(value || '').trim().replace(/\/+$/, '');
            const normalizeText = (value) => String(value || '').replace(/\s+/g, ' ').trim();
            const responseSample = normalizeText(responseText).slice(0, 120);

            const extractMessageText = (message) => {
                if (!message) {
                    return '';
                }
                const blocks = Array.isArray(message.content_block) ? message.content_block : [];
                for (const block of blocks) {
                    const text = block?.content?.text_block?.text;
                    if (text) {
                        return String(text).trim();
                    }
                }
                return String(message?.tts_content || message?.content || message?.brief || '').trim();
            };
            const messageMatchesResponse = (message) => {
                if (!responseSample) {
                    return true;
                }
                const messageText = normalizeText(extractMessageText(message));
                if (!messageText) {
                    return false;
                }
                const messageSample = messageText.slice(0, 120);
                return (
                    messageText.includes(responseSample) ||
                    responseSample.includes(messageSample) ||
                    messageSample.includes(responseSample)
                );
            };

            const getSearchMessage = (conversation) => {
                if (!conversation || !Array.isArray(conversation.messages)) {
                    return null;
                }

                const messages = conversation.messages;
                for (const message of messages) {
                    if (message?.user_type !== 2) {
                        continue;
                    }
                    if (!messageMatchesResponse(message)) {
                        continue;
                    }
                    const blocks = Array.isArray(message?.content_block) ? message.content_block : [];
                    if (blocks.some((block) => block?.content?.search_query_result_block)) {
                        return message;
                    }
                }

                if (promptText) {
                    const normalizedPrompt = normalizeText(promptText);
                    for (const message of messages) {
                        if (message?.user_type !== 2) {
                            continue;
                        }
                        const messageText = normalizeText(extractMessageText(message));
                        if (
                            messageText &&
                            (messageText.includes(normalizedPrompt) ||
                                normalizedPrompt.includes(messageText.slice(0, 80)))
                        ) {
                            const blocks = Array.isArray(message?.content_block) ? message.content_block : [];
                            if (blocks.some((block) => block?.content?.search_query_result_block)) {
                                return message;
                            }
                        }
                    }
                }

                return null;
            };

            const conversations = cells
                .map((cell) => cell?.conversation)
                .filter((conversation) => conversation && Array.isArray(conversation.messages));
            conversations.sort((left, right) => {
                const leftActive = String(left?.conversation_id || '') === activeConversationId ? 1 : 0;
                const rightActive = String(right?.conversation_id || '') === activeConversationId ? 1 : 0;
                if (leftActive !== rightActive) {
                    return rightActive - leftActive;
                }
                const leftTime = Number(left?.update_time || left?.create_time || 0);
                const rightTime = Number(right?.update_time || right?.create_time || 0);
                return rightTime - leftTime;
            });

            for (const conversation of conversations) {
                const message = getSearchMessage(conversation);
                if (!message) {
                    continue;
                }

                const blocks = Array.isArray(message?.content_block) ? message.content_block : [];
                const collected = [];

                for (const block of blocks) {
                    const searchBlock = block?.content?.search_query_result_block;
                    if (!searchBlock) {
                        continue;
                    }

                    const titleLookup = new Map();
                    const results = Array.isArray(searchBlock.results) ? searchBlock.results : [];
                    for (const result of results) {
                        const card = result?.text_card;
                        if (!card) {
                            continue;
                        }
                        const url = String(card.url || '').trim();
                        const title = String(card.title || '').trim();
                        const appName = String(card.sitename || '').trim();
                        const docId = String(card.doc_id || '').trim();
                        if (url) {
                            titleLookup.set(normalizeUrl(url), { title, appName });
                        }
                        if (docId) {
                            titleLookup.set(`doc:${docId}`, { title, appName });
                        }
                        if (url || title || appName) {
                            collected.push({ url, title, app_name: appName });
                        }
                    }

                    const mediaResults = Array.isArray(searchBlock.vlm_rich_media_results)
                        ? searchBlock.vlm_rich_media_results
                        : [];
                    for (const media of mediaResults) {
                        const source = media?.image || media?.video || media?.audio || {};
                        const url = String(source.main_site_url || source.host_page_url || '').trim();
                        const docId = String(source.doc_id || '').trim();
                        const fromUrl = titleLookup.get(normalizeUrl(url)) || null;
                        const fromDoc = titleLookup.get(`doc:${docId}`) || null;
                        const matched = fromUrl || fromDoc || {};
                        const title = String(
                            matched.title || source.video_captions || source.img_caption || ''
                        ).trim();
                        const appName = String(
                            source.source_app_name || matched.appName || ''
                        ).trim();
                        if (url || title || appName) {
                            collected.push({ url, title, app_name: appName });
                        }
                    }
                }

                if (collected.length > 0) {
                    return collected;
                }
            }

            return [];
        }"""
            ,
            {
                "expectedPrompt": expected_prompt,
                "expectedResponseText": expected_response_text,
            },
        )
    except PlaywrightError:
        return []
    if not isinstance(payload, list):
        return []

    records: list[SourceRecord] = []
    for item in payload:
        if not isinstance(item, dict):
            continue
        records.append(
            SourceRecord(
                url=str(item.get("url") or "").strip(),
                title=str(item.get("title") or "").strip(),
                app_name=str(item.get("app_name") or "").strip(),
            )
        )
    return dedupe_source_records(records)


def get_structured_source_records(
    page,
    expected_prompt: str = "",
    expected_response_text: str = "",
) -> list[SourceRecord]:
    current_url = ""
    try:
        current_url = page.url
    except PlaywrightError:
        current_url = ""

    if "/chat/" in current_url:
        temp_page = page.context.new_page()
        try:
            temp_page.goto(current_url, wait_until="domcontentloaded")
            temp_page.wait_for_timeout(2000)
            records = _extract_structured_source_records_from_loaded_page(
                temp_page,
                expected_prompt=expected_prompt,
                expected_response_text=expected_response_text,
            )
            if records:
                return records
        except PlaywrightError:
            pass
        finally:
            try:
                temp_page.close()
            except PlaywrightError:
                pass

    return _extract_structured_source_records_from_loaded_page(
        page,
        expected_prompt=expected_prompt,
        expected_response_text=expected_response_text,
    )


def get_last_response_data(
    page,
    config: AppConfig,
    selectors: list[str],
    expected_prompt: str = "",
    include_sources: bool = True,
) -> ResponseData:
    for selector in selectors:
        locator = page.locator(selector)
        count = locator.count()
        if count == 0:
            continue
        response = locator.nth(count - 1)
        try:
            payload = response.evaluate(
                """(el) => ({
                    text: (el.innerText || '').trim(),
                    links: Array.from(el.querySelectorAll('a[href]')).map((link) => ({
                        text: (link.innerText || '').trim(),
                        href: link.href || ''
                    }))
                })"""
            )
        except PlaywrightError:
            continue

        text = str(payload.get("text") or "").strip()
        if text:
            links = payload.get("links") or []
            if not include_sources:
                return ResponseData(text=text, sources="")
            structured_sources = get_structured_source_records(
                page,
                expected_prompt=expected_prompt,
                expected_response_text=text,
            )
            if structured_sources:
                sources, source_urls, source_titles = build_structured_source_fields(structured_sources)
                return ResponseData(
                    text=text,
                    sources=sources,
                    source_urls=source_urls,
                    source_titles=source_titles,
                )
            nearby_sources = get_nearby_dom_source_records(page, selectors)
            if nearby_sources:
                sources, source_urls, source_titles = build_structured_source_fields(nearby_sources)
                return ResponseData(
                    text=text,
                    sources=sources,
                    source_urls=source_urls,
                    source_titles=source_titles,
                )
            panel_sources = get_reference_panel_source_records(page)
            if panel_sources:
                sources, source_urls, source_titles = build_structured_source_fields(panel_sources)
                return ResponseData(
                    text=text,
                    sources=sources,
                    source_urls=source_urls,
                    source_titles=source_titles,
                )
            sources = format_sources(text, links)
            if not sources:
                try:
                    debug_path = save_source_debug_artifact(
                        page,
                        config,
                        expected_prompt=expected_prompt,
                        expected_response_text=text,
                        links=links,
                    )
                    print(f"Source debug artifact saved to {debug_path}")
                except Exception as exc:
                    print(f"Failed to save source debug artifact: {exc}")
            return ResponseData(text=text, sources=sources)
    return ResponseData(text="", sources="")


def get_response_count(page, selectors: list[str]) -> int:
    for selector in selectors:
        count = page.locator(selector).count()
        if count > 0:
            return count
    return 0


def is_loading(page, selectors: list[str]) -> bool:
    for selector in selectors:
        locator = page.locator(selector)
        if locator.count() == 0:
            continue
        try:
            if locator.first.is_visible():
                return True
        except PlaywrightError:
            continue
    return False


def prepare_input(locator, prompt: str, clear_input_hotkey: str) -> None:
    locator.click()
    try:
        locator.fill("")
        locator.fill(prompt)
        return
    except PlaywrightError:
        pass

    locator.press(clear_input_hotkey)
    locator.press("Backspace")
    locator.type(prompt, delay=20)


def wait_for_response(
    page,
    config: AppConfig,
    previous_count: int,
    previous_text: str,
    expected_prompt: str,
) -> str:
    deadline = time.time() + config.chat.response_timeout_seconds
    latest_text = previous_text
    stable_rounds = 0

    while time.time() < deadline:
        handle_popup_if_present(page, config)
        current_count = get_response_count(page, config.chat.response_selectors)
        current_response = get_last_response_data(
            page,
            config,
            config.chat.response_selectors,
            expected_prompt=expected_prompt,
            include_sources=False,
        )
        current_text = current_response.text
        response_arrived = current_count > previous_count or (
            current_text and current_text != previous_text
        )

        if response_arrived and current_text:
            if current_text == latest_text:
                stable_rounds += 1
            else:
                latest_text = current_text
                stable_rounds = 0

            if stable_rounds >= config.chat.stability_checks and not is_loading(
                page, config.chat.loading_selectors
            ):
                return get_last_response_data(
                    page,
                    config,
                    config.chat.response_selectors,
                    expected_prompt=expected_prompt,
                    include_sources=True,
                )

        time.sleep(config.chat.poll_interval_seconds)

    raise TimeoutError(f"Timed out waiting for {config.chat.platform_name} to finish responding.")


def send_prompt(page, config: AppConfig, prompt: str) -> ResponseData:
    if config.chat.new_chat_each_prompt:
        click_if_present(page, config.chat.new_chat_selectors, config.browser.action_timeout_ms)
        time.sleep(1)

    handle_popup_if_present(page, config)
    _, input_locator = ensure_chat_ready(page, config)
    previous_count = get_response_count(page, config.chat.response_selectors)
    previous_text = get_last_response_data(
        page,
        config,
        config.chat.response_selectors,
        include_sources=False,
    ).text

    prepare_input(input_locator, prompt, config.chat.clear_input_hotkey)
    if not click_if_present(page, config.chat.send_button_selectors, config.browser.action_timeout_ms):
        input_locator.press(config.chat.send_hotkey)

    handle_popup_if_present(page, config)
    return wait_for_response(page, config, previous_count, previous_text, prompt)


@contextmanager
def open_chat_page(config: AppConfig, interactive_login: bool = True):
    config.browser.user_data_dir.mkdir(parents=True, exist_ok=True)

    with sync_playwright() as playwright:
        launch_kwargs = {
            "user_data_dir": str(config.browser.user_data_dir),
            "headless": config.browser.headless,
        }
        if config.browser.channel:
            launch_kwargs["channel"] = config.browser.channel

        context = playwright.chromium.launch_persistent_context(**launch_kwargs)
        context.set_default_timeout(config.browser.action_timeout_ms)
        page = context.pages[0] if context.pages else context.new_page()
        page.goto(config.browser.start_url, wait_until="domcontentloaded")
        page.wait_for_timeout(config.browser.startup_wait_ms)

        if config.chat.manual_login:
            try:
                ensure_chat_ready(page, config)
            except RuntimeError:
                if not interactive_login:
                    context.close()
                    raise RuntimeError(
                        f"Login required for {config.chat.platform_name}. "
                        "Complete login in the configured browser profile first."
                    )
                print(
                    f"Please log in to {config.chat.platform_name} in the opened browser window, "
                    "then press Enter here."
                )
                input()
                ensure_chat_ready(page, config)

        handle_popup_if_present(page, config)
        try:
            yield context, page
        finally:
            context.close()


def run_prompt_batch(
    config: AppConfig,
    prompts: list[str],
    interactive_login: bool = True,
) -> list[PromptRunRecord]:
    records: list[PromptRunRecord] = []

    with open_chat_page(config, interactive_login=interactive_login) as (_, page):
        for prompt in prompts:
            prompt_text = str(prompt).strip()
            if not prompt_text:
                continue
            response = send_prompt(page, config, prompt_text)
            records.append(
                PromptRunRecord(
                    query=prompt_text,
                    result=response.text,
                    sources=response.sources,
                    source_urls=response.source_urls,
                    source_titles=response.source_titles,
                )
            )

    return records


def export_prompt_records_to_excel(
    records: list[PromptRunRecord],
    output_path: Path,
    summary: Optional[dict[str, str]] = None,
) -> Path:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Results"
    headers = ["query", "result", "sources", "source_urls", "source_titles"]
    sheet.append(headers)

    for record in records:
        sheet.append(
            [
                record.query,
                record.result,
                record.sources,
                record.source_urls,
                record.source_titles,
            ]
        )

    if summary:
        summary_sheet = workbook.create_sheet("Summary")
        summary_sheet.append(["field", "value"])
        for key, value in summary.items():
            summary_sheet.append([key, value])

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    return output_path


def main() -> int:
    args = parse_args()
    config_path = Path(args.config).expanduser().resolve()
    config = load_config(config_path)
    if not config.excel.path.exists():
        raise FileNotFoundError(
            f"Excel file not found: {config.excel.path}. "
            "Update excel.path in the config or create the workbook first."
        )

    workbook = load_workbook(config.excel.path)
    sheet = workbook[config.excel.sheet] if config.excel.sheet else workbook.active
    prompt_col = column_index_by_header(sheet, config.excel.header_row, config.excel.prompt_column)
    result_col = column_index_by_header(sheet, config.excel.header_row, config.excel.result_column)
    source_col = ensure_column_index_by_header(sheet, config.excel.header_row, config.excel.source_column)
    source_urls_col = ensure_column_index_by_header(
        sheet,
        config.excel.header_row,
        "source_urls" if source_col is not None else None,
    )
    source_titles_col = ensure_column_index_by_header(
        sheet,
        config.excel.header_row,
        "source_titles" if source_col is not None else None,
    )

    rows = list(iter_candidate_rows(sheet, config.excel, overwrite=args.overwrite))
    if args.limit is not None:
        rows = rows[: args.limit]

    if not rows:
        print("No eligible rows found. Nothing to do.")
        return 0

    with open_chat_page(config, interactive_login=True) as (_, page):
        for index, row_idx in enumerate(rows, start=1):
            prompt = str(sheet.cell(row=row_idx, column=prompt_col).value).strip()
            print(f"[{index}/{len(rows)}] Processing row {row_idx}")
            try:
                response = send_prompt(page, config, prompt)
                result = response.text
                sources = response.sources
                source_urls = response.source_urls
                source_titles = response.source_titles
            except (PlaywrightError, PlaywrightTimeoutError, TimeoutError) as exc:
                result = f"ERROR: {exc}"
                sources = ""
                source_urls = ""
                source_titles = ""
            sheet.cell(row=row_idx, column=result_col).value = result
            if source_col is not None:
                sheet.cell(row=row_idx, column=source_col).value = sources
            if source_urls_col is not None:
                sheet.cell(row=row_idx, column=source_urls_col).value = source_urls
            if source_titles_col is not None:
                sheet.cell(row=row_idx, column=source_titles_col).value = source_titles
            workbook.save(config.excel.path)

    print(f"Finished. Results written to {config.excel.path}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
